"""Шаблон Excel для расчёта КЛ свыше 1 кВ (проверка на невозгорание)."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

from app.shared.excel_project_base import require_openpyxl

from .defaults import demo_mv_cable_input

SHEET = "расчёт"

META_U_N = "Uн, кВ"
META_RU_TYPE = "Тип РУ"
META_PROJECT = "Проект"
META_SUBSTATION = "Подстанция"
META_T_BREAKER = "t выключателя, с"
META_Q0 = "Q0, °C"
META_Q_OKR = "Qос, °C"
META_IKZ3 = "Iкз(3), кА"

ROW_HEADERS: tuple[str, ...] = (
    "№ кабеля",
    "Тип линии",
    "P, кВА",
    "Iн, А",
    "Нов/Сущ.",
    "Марка кабеля",
    "Количество жил",
    "Сечение, мм²",
    "Материал жил",
    "Материал оболочки, р.п.",
    "Iдд, А",
    "Qдд, °C",
    "tоткл., с (основная защита)",
    "tоткл., с (резервная защита)",
    "Та.эк., с",
    "b, мм2/кА2·с",
    "Tдоп нагрев, °C",
    "Tдоп невозгорание, °C",
)

TEMPLATE_FILENAME = "Шаблон КЛ свыше 1 кВ.xlsx"

_HEADER_ROW = 11
_EXAMPLE_ROW = _HEADER_ROW + 1


def _column_width(title: str, *, minimum: float = 8.0, padding: float = 2.0) -> float:
    longest = max((len(line) for line in title.splitlines()), default=len(title))
    return max(minimum, longest + padding)


def _example_row_values() -> tuple[object, ...]:
    m = demo_mv_cable_input()
    return (
        1,
        m.load_name,
        m.tsn_power_kva,
        "",
        "Нов",
        m.cable_mark,
        m.cores_count,
        m.section_mm2,
        m.conductor_material,
        m.sheath_material_gen,
        m.i_dd_a,
        m.theta_dd_c,
        m.t_main_protection_s,
        m.t_backup_protection_s,
        m.tae_s,
        m.b_const,
        m.theta_limit_heating_c,
        m.theta_limit_fire_c,
    )


def build_mv_cable_excel_template_bytes() -> bytes:
    openpyxl = require_openpyxl()
    from openpyxl.utils import get_column_letter

    demo = demo_mv_cable_input()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET

    ws["A1"] = "Поле"
    ws["B1"] = "Значение"
    meta_rows: tuple[tuple[str, object], ...] = (
        (META_PROJECT, demo.project_name or "Выбор и проверка КЛ 6 кВ до ТСН"),
        (META_U_N, demo.voltage_kv),
        (META_RU_TYPE, "ЗРУ"),
        (META_SUBSTATION, demo.substation_name),
        (META_IKZ3, demo.ikz3_ka),
        (META_T_BREAKER, demo.t_breaker_s),
        (META_Q0, demo.theta_0_c),
        (META_Q_OKR, demo.theta_okr_c),
    )
    for idx, (key, value) in enumerate(meta_rows, start=2):
        ws.cell(row=idx, column=1, value=key)
        ws.cell(row=idx, column=2, value=value)

    _a_labels = ["Поле", *(key for key, _ in meta_rows)]
    _meta_a_width = max(_column_width(label, minimum=12.0) for label in _a_labels)
    _meta_b_width = _column_width("Значение", minimum=14.0)

    for col, title in enumerate(ROW_HEADERS, start=1):
        ws.cell(row=_HEADER_ROW, column=col, value=title)

    for col, value in enumerate(_example_row_values(), start=1):
        ws.cell(row=_EXAMPLE_ROW, column=col, value=value)

    for col, title in enumerate(ROW_HEADERS, start=1):
        letter = get_column_letter(col)
        width = _column_width(title)
        if col == 1:
            width = max(width, _meta_a_width)
        elif col == 2:
            width = max(width, _meta_b_width)
        ws.column_dimensions[letter].width = width

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def save_mv_cable_excel_template(path: Path) -> Path:
    target = Path(path)
    if target.suffix.lower() != ".xlsx":
        target = target.with_suffix(".xlsx")
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(build_mv_cable_excel_template_bytes())
    return target
